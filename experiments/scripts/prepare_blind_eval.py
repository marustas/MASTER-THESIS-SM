"""
Blind expert-evaluation prep — supervisor revision (2026-06-07).

Generates a bias-controlled labelling pack:
  - labelling.xlsx   — single sheet, 45 rows × (1 + 10 archetype + 30 gap) cols.
                       Anonymised programme IDs, dropdown labels only,
                       NO institution, NO programme name, NO matching scores.
  - descriptions.xlsx — reference file with programme descriptions and
                       job-archetype descriptions.
  - internal_key.csv — NOT for the labeller; maps anon_id back to the real
                       programme id, name, institution, primary bucket
                       (ambiguous/short_desc/sparse_skills/clean).

Design (supervisor brief):
  * Mix three description-quality buckets randomly so a labeller cannot infer
    the bucket from the description style; randomised order also controls for
    labeller fatigue.
  * Uniform job-archetype columns across all programmes — match labels
    are comparable per row.
  * Cells use 4-option data validation: agree / disagree / neutral / can't judge.
  * Per-programme gap-skill list is accumulated over the algorithm's top-10
    matched jobs (∪ job skills minus programme skills, ranked by demand
    frequency, capped at 15).
  * Labeller never sees: real name, institution, bucket, ranking scores,
    job matches, or the algorithm's gap derivation.
"""

from __future__ import annotations

import json
import random
from collections import Counter
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Paths ─────────────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parents[2]
DATASET = ROOT / "data" / "dataset" / "dataset.parquet"
RANKINGS = ROOT / "experiments" / "results" / "exp3_hybrid" / "rankings.parquet"
OUT_DIR = ROOT / "experiments" / "expert_eval" / "blind_eval"
OUT_DIR.mkdir(parents=True, exist_ok=True)

LABELLING_PATH = OUT_DIR / "labelling.xlsx"
DESCRIPTIONS_PATH = OUT_DIR / "descriptions.xlsx"
INTERNAL_KEY_PATH = OUT_DIR / "internal_key.csv"

# Deterministic shuffle so re-runs produce the same anon_id mapping.
SEED = 20260607

TOP_K_FOR_GAPS = 10        # algorithm's top-K used to accumulate gap skills
MAX_GAP_SLOTS = 15         # cap of gap skills shown per programme
LABEL_OPTIONS = ["agree", "disagree", "neutral", "can't judge"]

# Description-quality slice size (programmes per bucket) and priority order
# for primary-tag assignment when a programme is in several.
BUCKET_SLICE_SIZE = 15
BUCKET_PRIORITY = ["ambiguous_match", "sparse_skills", "short_desc"]


# ── Job archetypes (uniform across all programmes) ────────────────────────────
# Curated to span the ICT field. Header columns in the labelling sheet are the
# archetype names; descriptions live in descriptions.xlsx.
JOB_ARCHETYPES: list[tuple[str, str]] = [
    (
        "Software Developer (Generalist)",
        "Designs, codes and tests application software in mainstream languages "
        "(Python, Java, C#, JavaScript). Works across web, backend and "
        "general-purpose business applications. Uses Git, agile delivery, and "
        "standard frameworks. Entry-level expectation: write clean code from "
        "specifications, debug, contribute to small features under supervision.",
    ),
    (
        "Data Analyst / BI",
        "Extracts, cleans and analyses business data; builds reports and "
        "dashboards (Power BI, Tableau, Looker). Strong SQL, basic Python or R, "
        "statistical literacy. Translates business questions into data queries "
        "and visualisations for non-technical stakeholders.",
    ),
    (
        "Data Engineer / Database Developer",
        "Designs and maintains data pipelines, ETL jobs and database schemas. "
        "Works with relational (PostgreSQL, MS SQL) and increasingly cloud "
        "warehouses (Snowflake, BigQuery). Writes performant SQL and pipeline "
        "code; ensures data quality, lineage and availability.",
    ),
    (
        "IT Support / Service Desk",
        "First and second-line user support: incident triage, ticketing, "
        "hardware/software troubleshooting, account and access management. "
        "Communicates with end-users, follows ITIL-style processes. Entry-level "
        "ICT role, strong customer-service orientation.",
    ),
    (
        "Systems Administrator / IT Infrastructure",
        "Operates and maintains server and OS infrastructure (Windows Server, "
        "Linux), virtualisation, backups, Active Directory, monitoring. "
        "Increasingly hybrid with cloud (Azure, AWS basics). Responsible for "
        "uptime, patching and operational documentation.",
    ),
    (
        "Cybersecurity Analyst",
        "Monitors security events (SIEM), investigates incidents, runs "
        "vulnerability scans, supports access reviews and security awareness. "
        "Understands the OWASP top 10, network protocols, common attack "
        "patterns. Often SOC L1/L2 or junior GRC.",
    ),
    (
        "Network / DevOps Engineer",
        "Configures networks (routing, switching, firewalls) and / or CI/CD "
        "pipelines, container orchestration (Docker, Kubernetes), "
        "infrastructure-as-code (Terraform, Ansible). Bridges development and "
        "operations; automates deployment and observability.",
    ),
    (
        "QA / Test Engineer",
        "Designs test cases, executes manual and automated tests, files defects. "
        "Familiar with test frameworks (Selenium, Playwright, Cypress, pytest), "
        "API testing (Postman), basic SQL. Junior roles emphasise structured "
        "thinking and attention to detail over deep coding.",
    ),
    (
        "Project / Product Manager (IT)",
        "Coordinates ICT projects or product backlogs: scope, timeline, "
        "stakeholders, risk. Works in agile (Scrum, Kanban) or hybrid models. "
        "Junior variants include business analyst / scrum master / associate PM. "
        "Strong communication, requirements analysis, basic technical literacy.",
    ),
    (
        "AI / ML Engineer",
        "Builds and deploys machine-learning models for production use: data "
        "preparation, model training (scikit-learn, PyTorch, TensorFlow), "
        "evaluation, MLOps. Strong Python; growing demand for LLM / RAG "
        "integration and generative-AI tooling.",
    ),
]
ASSERTION_HEADER = "Statement to label: 'This programme prepares graduates for the role.'"

# ── Helpers ───────────────────────────────────────────────────────────────────


def _primary_bucket(prog_id: int, bucket_membership: dict[str, set[int]]) -> str:
    for b in BUCKET_PRIORITY:
        if prog_id in bucket_membership[b]:
            return b
    return "clean"


def _safe_len(value) -> int:
    if isinstance(value, (list, tuple, set, np.ndarray)):
        return len(value)
    return 0


def _compute_bucket_membership(
    programmes: pd.DataFrame,
    rankings: pd.DataFrame,
) -> dict[str, set[int]]:
    """Derive bucket membership directly from the dataset.

    Three description-quality slices, each containing the bottom-15 programmes:
      - ``short_desc``      — shortest ``cleaned_text``
      - ``sparse_skills``   — fewest distinct extracted ESCO URIs
      - ``ambiguous_match`` — smallest top-1 vs top-2 ``hybrid_score`` gap

    ``programmes`` must be the programmes-only subset with positional index
    matching the ``programme_id`` convention used by ``rankings.parquet``.
    """
    desc_len = programmes["cleaned_text"].fillna("").astype(str).str.len()
    short = set(int(i) for i in desc_len.nsmallest(BUCKET_SLICE_SIZE).index)

    n_skills = programmes["skill_uris"].apply(_safe_len)
    sparse = set(int(i) for i in n_skills.nsmallest(BUCKET_SLICE_SIZE).index)

    gaps: list[tuple[int, float]] = []
    for pid, g in rankings.groupby("programme_id"):
        top2 = g.nlargest(2, "hybrid_score")
        if len(top2) >= 2:
            gaps.append(
                (int(pid), float(top2.iloc[0]["hybrid_score"] - top2.iloc[1]["hybrid_score"]))
            )
    gaps.sort(key=lambda t: t[1])
    ambiguous = {pid for pid, _ in gaps[:BUCKET_SLICE_SIZE]}

    return {
        "short_desc":      short,
        "sparse_skills":   sparse,
        "ambiguous_match": ambiguous,
    }


_DESC_CHAR_CAP = 4000  # truncate over-long descriptions for labeller ergonomics


def _description_for(row: pd.Series) -> str:
    """Return the description text the algorithm actually consumed.

    ``cleaned_text`` is what the embedding model and skill extractor saw, so
    showing it to the labeller keeps the human judgement aligned with the
    algorithm's information. ``extended_description`` is the raw scrape, which
    frequently contains nav-menu boilerplate ('ABOUT AIKOSNEWSREGISTERS …')
    that the cleaner removes; using it would tip the labeller off about
    extraction quality before they read the substance.
    """
    for col in ("cleaned_text", "extended_description", "brief_description"):
        v = row.get(col)
        if isinstance(v, str) and v.strip() and v.strip().lower() != "none":
            text = v.strip()
            if len(text) > _DESC_CHAR_CAP:
                text = text[:_DESC_CHAR_CAP].rstrip() + " […truncated]"
            return text
    return ""


def _skill_uris(row: pd.Series) -> set[str]:
    details = row.get("skill_details", [])
    if not isinstance(details, (list, tuple, np.ndarray)):
        return set()
    return {d.get("esco_uri") for d in details if d.get("esco_uri")}


def _uri_to_label(df: pd.DataFrame) -> dict[str, str]:
    """Pull a preferred label for each ESCO URI from any document that mentions it."""
    mapping: dict[str, str] = {}
    for _, row in df.iterrows():
        details = row.get("skill_details", [])
        if not isinstance(details, (list, tuple, np.ndarray)):
            continue
        for d in details:
            uri = d.get("esco_uri")
            lbl = d.get("preferred_label")
            if uri and lbl and uri not in mapping:
                mapping[uri] = lbl
    return mapping


def _compute_gap_skills(
    top_job_ids: list[int],
    job_skills_by_id: dict[int, set[str]],
    prog_skills: set[str],
    uri_label: dict[str, str],
) -> list[str]:
    """Aggregate skill demand across the top-K matched jobs, subtract
    programme skills. ``top_job_ids`` and ``job_skills_by_id`` keys must be
    the global ``dataset.parquet`` row indices.
    """
    demand: Counter[str] = Counter()
    for jid in top_job_ids:
        for u in job_skills_by_id.get(jid, set()):
            demand[u] += 1
    missing = [(u, c) for u, c in demand.items() if u not in prog_skills]
    missing.sort(key=lambda t: (-t[1], uri_label.get(t[0], t[0])))
    return [uri_label.get(u, u) for u, _ in missing[:MAX_GAP_SLOTS]]


# ── Build ─────────────────────────────────────────────────────────────────────


def build() -> dict:
    random.seed(SEED)
    np.random.seed(SEED)

    df = pd.read_parquet(DATASET)
    rankings = pd.read_parquet(RANKINGS)

    # rankings.programme_id / job_id are the global ``dataset.parquet`` row
    # indices.  Keep df indices throughout to align bucket lookups, top-K
    # candidate joins and skill-set lookups in a single coordinate system.
    programmes = df[df["source_type"] == "programme"]
    jobs = df[df["source_type"] == "job_ad"]
    uri_label = _uri_to_label(df)

    job_skills_by_id: dict[int, set[str]] = {
        int(idx): _skill_uris(row) for idx, row in jobs.iterrows()
    }

    bucket_membership = _compute_bucket_membership(programmes, rankings)

    # Top-K matched job df indices per programme df index.
    top_k_job_ids = (
        rankings.sort_values(["programme_id", "hybrid_score"], ascending=[True, False])
        .groupby("programme_id", sort=False)
        .head(TOP_K_FOR_GAPS)
        .groupby("programme_id")["job_id"]
        .apply(lambda s: [int(x) for x in s])
        .to_dict()
    )

    records: list[dict] = []
    for idx, row in programmes.iterrows():
        prog_id = int(idx)
        prog_uris = _skill_uris(row)
        gap_skills = _compute_gap_skills(
            top_k_job_ids.get(prog_id, []),
            job_skills_by_id,
            prog_uris,
            uri_label,
        )
        records.append({
            "real_programme_id": prog_id,
            "name": str(row["name"]),
            "institution": str(row.get("institution", "")),
            "primary_bucket": _primary_bucket(prog_id, bucket_membership),
            "description": _description_for(row),
            "gap_skills": gap_skills,
        })

    # Random shuffle → anon_id assignment.
    order = list(range(len(records)))
    random.shuffle(order)
    anon_records = []
    for new_pos, original_idx in enumerate(order, start=1):
        rec = dict(records[original_idx])
        rec["anon_id"] = f"P{new_pos:02d}"
        anon_records.append(rec)

    _write_labelling(anon_records)
    _write_descriptions(anon_records)
    _write_internal_key(anon_records)

    stats = {
        "n_programmes": len(anon_records),
        "buckets": Counter(r["primary_bucket"] for r in anon_records),
        "gap_skill_counts": Counter(len(r["gap_skills"]) for r in anon_records),
        "n_archetypes": len(JOB_ARCHETYPES),
        "label_options": LABEL_OPTIONS,
        "outputs": {
            "labelling": str(LABELLING_PATH),
            "descriptions": str(DESCRIPTIONS_PATH),
            "internal_key": str(INTERNAL_KEY_PATH),
        },
    }
    return stats


# ── Excel writers ─────────────────────────────────────────────────────────────


_HEADER_FONT = Font(bold=True)
_HEADER_FILL = PatternFill("solid", fgColor="DCE6F1")
_GAP_HEADER_FILL = PatternFill("solid", fgColor="FFF2CC")
_SKILL_NAME_FILL = PatternFill("solid", fgColor="F2F2F2")
_WRAP = Alignment(wrap_text=True, vertical="top")


def _write_labelling(records: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Labelling"

    # ── Headers ───────────────────────────────────────────────────────────────
    headers = ["anon_id"]
    headers.extend(name for name, _desc in JOB_ARCHETYPES)
    for i in range(1, MAX_GAP_SLOTS + 1):
        headers.append(f"gap_skill_{i:02d}")
        headers.append(f"gap_label_{i:02d}")
    ws.append(headers)
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.alignment = _WRAP
        if h.startswith("gap_"):
            cell.fill = _GAP_HEADER_FILL
        else:
            cell.fill = _HEADER_FILL

    # ── Data validation: archetype + gap-label columns ────────────────────────
    formula = '"' + ",".join(LABEL_OPTIONS) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "Pick one of: " + ", ".join(LABEL_OPTIONS)
    dv.errorTitle = "Invalid label"
    ws.add_data_validation(dv)

    # ── Rows ──────────────────────────────────────────────────────────────────
    for row_idx, rec in enumerate(records, start=2):
        ws.cell(row=row_idx, column=1, value=rec["anon_id"]).font = _HEADER_FONT
        # Archetype dropdown cells (cols 2..1+len(JOB_ARCHETYPES))
        last_arch_col = 1 + len(JOB_ARCHETYPES)
        for c in range(2, last_arch_col + 1):
            dv.add(ws.cell(row=row_idx, column=c))

        # Gap pairs (skill_name | label) — skill_name pre-filled, read-only by convention
        gap_skills = rec["gap_skills"]
        first_gap_col = last_arch_col + 1
        for i in range(MAX_GAP_SLOTS):
            name_col = first_gap_col + 2 * i
            label_col = name_col + 1
            if i < len(gap_skills):
                name_cell = ws.cell(row=row_idx, column=name_col, value=gap_skills[i])
                name_cell.fill = _SKILL_NAME_FILL
                name_cell.alignment = _WRAP
                dv.add(ws.cell(row=row_idx, column=label_col))

    # ── Column widths & freeze ────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 8
    for c in range(2, 1 + len(JOB_ARCHETYPES) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 22
    for i in range(MAX_GAP_SLOTS):
        name_col = 1 + len(JOB_ARCHETYPES) + 1 + 2 * i
        label_col = name_col + 1
        ws.column_dimensions[get_column_letter(name_col)].width = 30
        ws.column_dimensions[get_column_letter(label_col)].width = 14
    ws.freeze_panes = "B2"
    ws.row_dimensions[1].height = 60

    wb.save(LABELLING_PATH)


def _write_descriptions(records: list[dict]) -> None:
    wb = Workbook()

    # ── Sheet 1: Programmes ───────────────────────────────────────────────────
    ws_p = wb.active
    ws_p.title = "Programmes"
    ws_p.append(["anon_id", "description"])
    for cell in ws_p[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    for rec in records:
        ws_p.append([rec["anon_id"], rec["description"]])
    ws_p.column_dimensions["A"].width = 10
    ws_p.column_dimensions["B"].width = 140
    for row in ws_p.iter_rows(min_row=2):
        for c in row:
            c.alignment = _WRAP

    # ── Sheet 2: Job archetypes ───────────────────────────────────────────────
    ws_j = wb.create_sheet("Job Archetypes")
    ws_j.append([ASSERTION_HEADER])
    ws_j["A1"].font = Font(bold=True, italic=True)
    ws_j["A1"].fill = _HEADER_FILL
    ws_j.append([])
    ws_j.append(["archetype", "description"])
    for cell in ws_j[3]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    for name, desc in JOB_ARCHETYPES:
        ws_j.append([name, desc])
    ws_j.column_dimensions["A"].width = 38
    ws_j.column_dimensions["B"].width = 110
    for row in ws_j.iter_rows(min_row=4):
        for c in row:
            c.alignment = _WRAP

    wb.save(DESCRIPTIONS_PATH)


def _write_internal_key(records: list[dict]) -> None:
    rows = [{
        "anon_id": r["anon_id"],
        "real_programme_id": r["real_programme_id"],
        "name": r["name"],
        "institution": r["institution"],
        "primary_bucket": r["primary_bucket"],
        "n_gap_skills": len(r["gap_skills"]),
    } for r in records]
    pd.DataFrame(rows).to_csv(INTERNAL_KEY_PATH, index=False)


if __name__ == "__main__":
    stats = build()
    print(json.dumps({
        "n_programmes": stats["n_programmes"],
        "buckets": dict(stats["buckets"]),
        "gap_skill_count_distribution": dict(sorted(stats["gap_skill_counts"].items())),
        "outputs": stats["outputs"],
    }, indent=2))
